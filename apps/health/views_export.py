"""
Health Export Views - Excel workout dashboard export.
"""

import logging
from collections import defaultdict
from datetime import timedelta
from io import BytesIO

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Avg, Count, Max, Sum
from django.http import HttpResponse
from django.views import View

from apps.core.utils import get_user_today

from .models import (
    CardioDetails,
    ClassDetails,
    Exercise,
    ExerciseSet,
    PersonalRecord,
    StepsEntry,
    WorkoutExercise,
    WorkoutSession,
)

logger = logging.getLogger(__name__)


class WorkoutDashboardExcelView(LoginRequiredMixin, View):
    """Export a formatted Excel workbook with a workout dashboard."""

    def get(self, request):
        import openpyxl
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        user = request.user
        today = get_user_today(user)
        wb = openpyxl.Workbook()

        # Style constants
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        stat_value_font = Font(bold=True, size=14)
        stat_label_font = Font(size=10, color="666666")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def style_header_row(ws, row, max_col):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        def style_data_cell(ws, row, col, number_format=None):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if number_format:
                cell.number_format = number_format
            return cell

        # ── Sheet 1: Dashboard Summary ──
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        ws_dash.sheet_properties.tabColor = "2B5797"

        workouts = WorkoutSession.objects.filter(user=user).order_by("-date")
        all_sets = ExerciseSet.objects.filter(
            workout_exercise__session__user=user,
        )
        steps = StepsEntry.objects.filter(user=user).order_by("-logged_date")
        prs = PersonalRecord.objects.filter(user=user).select_related("exercise").order_by("-achieved_date")

        # Title
        ws_dash.merge_cells("A1:F1")
        title_cell = ws_dash["A1"]
        title_cell.value = "Workout Dashboard"
        title_cell.font = Font(bold=True, size=18, color="2B5797")
        title_cell.alignment = Alignment(horizontal="center")

        ws_dash.merge_cells("A2:F2")
        ws_dash["A2"].value = f"Generated for {user.get_full_name() or user.email} — {today.strftime('%B %d, %Y')}"
        ws_dash["A2"].font = Font(size=10, color="888888")
        ws_dash["A2"].alignment = Alignment(horizontal="center")

        # Overview Stats
        row = 4
        ws_dash.merge_cells(f"A{row}:F{row}")
        ws_dash.cell(row=row, column=1, value="Overview").font = Font(bold=True, size=14, color="2B5797")
        row += 1

        total_workouts = workouts.count()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        workouts_this_week = workouts.filter(date__gte=week_ago).count()
        workouts_this_month = workouts.filter(date__gte=month_ago).count()
        total_duration = workouts.aggregate(s=Sum("duration_minutes"))["s"] or 0
        total_calories = workouts.aggregate(s=Sum("calories_burned"))["s"] or 0
        total_sets_count = all_sets.count()

        stat_labels = ["Total Workouts", "This Week", "This Month", "Total Minutes", "Total Calories", "Total Sets"]
        stat_values = [total_workouts, workouts_this_week, workouts_this_month, total_duration, total_calories, total_sets_count]

        for col_idx, (label, value) in enumerate(zip(stat_labels, stat_values), start=1):
            c = ws_dash.cell(row=row, column=col_idx, value=value)
            c.font = stat_value_font
            c.alignment = Alignment(horizontal="center")
            c = ws_dash.cell(row=row + 1, column=col_idx, value=label)
            c.font = stat_label_font
            c.alignment = Alignment(horizontal="center")

        # Workout type breakdown
        row += 3
        ws_dash.merge_cells(f"A{row}:C{row}")
        ws_dash.cell(row=row, column=1, value="Workouts by Type").font = Font(bold=True, size=14, color="2B5797")
        row += 1

        headers = ["Workout Type", "Count", "Avg Duration (min)"]
        for col_idx, h in enumerate(headers, start=1):
            ws_dash.cell(row=row, column=col_idx, value=h)
        style_header_row(ws_dash, row, 3)
        row += 1

        type_stats = (
            workouts.values("workout_type")
            .annotate(count=Count("id"), avg_dur=Avg("duration_minutes"))
            .order_by("-count")
        )
        type_chart_start = row
        for stat in type_stats:
            ws_dash.cell(row=row, column=1, value=stat["workout_type"] or "Unspecified").border = thin_border
            ws_dash.cell(row=row, column=2, value=stat["count"]).border = thin_border
            c = ws_dash.cell(row=row, column=3, value=round(stat["avg_dur"] or 0, 1))
            c.border = thin_border
            c.number_format = "0.0"
            row += 1
        type_chart_end = row - 1

        # Chart: Workouts by type
        if type_chart_end >= type_chart_start:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Workouts by Type"
            chart.y_axis.title = "Count"
            chart.style = 10
            data = Reference(ws_dash, min_col=2, min_row=type_chart_start - 1, max_row=type_chart_end)
            cats = Reference(ws_dash, min_col=1, min_row=type_chart_start, max_row=type_chart_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws_dash.add_chart(chart, f"E{type_chart_start - 1}")

        # Weekly volume (last 12 weeks)
        row += 1
        ws_dash.merge_cells(f"A{row}:D{row}")
        ws_dash.cell(row=row, column=1, value="Weekly Volume (Last 12 Weeks)").font = Font(bold=True, size=14, color="2B5797")
        row += 1

        headers = ["Week Starting", "Workouts", "Total Minutes", "Total Calories"]
        for col_idx, h in enumerate(headers, start=1):
            ws_dash.cell(row=row, column=col_idx, value=h)
        style_header_row(ws_dash, row, 4)
        row += 1

        weekly_chart_start = row
        for weeks_back in range(11, -1, -1):
            week_start = today - timedelta(days=today.weekday()) - timedelta(weeks=weeks_back)
            week_end = week_start + timedelta(days=6)
            week_workouts = workouts.filter(date__gte=week_start, date__lte=week_end)
            agg = week_workouts.aggregate(
                count=Count("id"),
                mins=Sum("duration_minutes"),
                cals=Sum("calories_burned"),
            )
            ws_dash.cell(row=row, column=1, value=week_start).border = thin_border
            ws_dash.cell(row=row, column=1).number_format = "MMM DD"
            ws_dash.cell(row=row, column=2, value=agg["count"]).border = thin_border
            ws_dash.cell(row=row, column=3, value=agg["mins"] or 0).border = thin_border
            ws_dash.cell(row=row, column=4, value=agg["cals"] or 0).border = thin_border
            row += 1
        weekly_chart_end = row - 1

        # Line chart: weekly workouts
        if weekly_chart_end >= weekly_chart_start:
            chart2 = LineChart()
            chart2.title = "Weekly Training Volume"
            chart2.y_axis.title = "Workouts"
            chart2.style = 10
            data = Reference(ws_dash, min_col=2, min_row=weekly_chart_start - 1, max_row=weekly_chart_end)
            cats = Reference(ws_dash, min_col=1, min_row=weekly_chart_start, max_row=weekly_chart_end)
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            ws_dash.add_chart(chart2, f"F{weekly_chart_start - 1}")

        # Set column widths
        for col, width in [(1, 20), (2, 15), (3, 18), (4, 18), (5, 15), (6, 15)]:
            ws_dash.column_dimensions[get_column_letter(col)].width = width

        # ── Sheet 2: Workout History ──
        ws_hist = wb.create_sheet("Workout History")
        ws_hist.sheet_properties.tabColor = "4CAF50"

        headers = ["Date", "Name", "Type", "Duration (min)", "Calories", "Exercises", "Total Sets", "Total Volume (lbs)", "Source"]
        for col_idx, h in enumerate(headers, start=1):
            ws_hist.cell(row=1, column=col_idx, value=h)
        style_header_row(ws_hist, 1, len(headers))

        r = 2
        for w in workouts.prefetch_related("workout_exercises__sets", "workout_exercises__exercise"):
            exercises = w.workout_exercises.all()
            ex_count = exercises.count()
            total_s = sum(we.sets.count() for we in exercises)
            total_v = sum(
                (s.weight or 0) * (s.reps or 0)
                for we in exercises
                for s in we.sets.all()
            )
            ws_hist.cell(row=r, column=1, value=w.date).border = thin_border
            ws_hist.cell(row=r, column=1).number_format = "YYYY-MM-DD"
            ws_hist.cell(row=r, column=2, value=w.name or "").border = thin_border
            ws_hist.cell(row=r, column=3, value=w.workout_type or "").border = thin_border
            ws_hist.cell(row=r, column=4, value=w.duration_minutes or "").border = thin_border
            ws_hist.cell(row=r, column=5, value=w.calories_burned or "").border = thin_border
            ws_hist.cell(row=r, column=6, value=ex_count).border = thin_border
            ws_hist.cell(row=r, column=7, value=total_s).border = thin_border
            c = ws_hist.cell(row=r, column=8, value=float(total_v) if total_v else "")
            c.border = thin_border
            if total_v:
                c.number_format = "#,##0"
            ws_hist.cell(row=r, column=9, value=w.source or "manual").border = thin_border
            r += 1

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws_hist.column_dimensions[get_column_letter(col)].width = max(14, len(headers[col - 1]) + 4)

        # ── Sheet 3: Exercise Details ──
        ws_ex = wb.create_sheet("Exercise Details")
        ws_ex.sheet_properties.tabColor = "FF9800"

        headers = ["Date", "Workout", "Exercise", "Category", "Set #", "Weight (lbs)", "Reps", "Volume", "Warmup?", "PR?"]
        for col_idx, h in enumerate(headers, start=1):
            ws_ex.cell(row=1, column=col_idx, value=h)
        style_header_row(ws_ex, 1, len(headers))

        r = 2
        workout_exercises = (
            WorkoutExercise.objects.filter(session__user=user)
            .select_related("session", "exercise")
            .prefetch_related("sets")
            .order_by("-session__date", "order")
        )
        for we in workout_exercises:
            for s in we.sets.all().order_by("set_number"):
                ws_ex.cell(row=r, column=1, value=we.session.date).border = thin_border
                ws_ex.cell(row=r, column=1).number_format = "YYYY-MM-DD"
                ws_ex.cell(row=r, column=2, value=we.session.name or "").border = thin_border
                ws_ex.cell(row=r, column=3, value=we.exercise.name).border = thin_border
                ws_ex.cell(row=r, column=4, value=we.exercise.category).border = thin_border
                ws_ex.cell(row=r, column=5, value=s.set_number).border = thin_border
                c = ws_ex.cell(row=r, column=6, value=float(s.weight) if s.weight else "")
                c.border = thin_border
                ws_ex.cell(row=r, column=7, value=s.reps or "").border = thin_border
                vol = float((s.weight or 0) * (s.reps or 0))
                c = ws_ex.cell(row=r, column=8, value=vol if vol else "")
                c.border = thin_border
                if vol:
                    c.number_format = "#,##0"
                ws_ex.cell(row=r, column=9, value="Yes" if s.is_warmup else "").border = thin_border
                ws_ex.cell(row=r, column=10, value="Yes" if s.is_pr else "").border = thin_border
                r += 1

        for col in range(1, len(headers) + 1):
            ws_ex.column_dimensions[get_column_letter(col)].width = max(12, len(headers[col - 1]) + 4)

        # ── Sheet 4: Cardio & Classes ──
        ws_cardio = wb.create_sheet("Cardio & Classes")
        ws_cardio.sheet_properties.tabColor = "E91E63"

        headers = ["Date", "Workout", "Exercise", "Type", "Duration (min)", "Distance (mi)", "Intensity", "Calories", "Avg HR"]
        for col_idx, h in enumerate(headers, start=1):
            ws_cardio.cell(row=1, column=col_idx, value=h)
        style_header_row(ws_cardio, 1, len(headers))

        r = 2
        cardio_exercises = (
            WorkoutExercise.objects.filter(
                session__user=user,
                exercise__category__in=["cardio", "class"],
            )
            .select_related("session", "exercise")
            .prefetch_related("cardio_details", "class_details")
            .order_by("-session__date")
        )
        for we in cardio_exercises:
            ws_cardio.cell(row=r, column=1, value=we.session.date).border = thin_border
            ws_cardio.cell(row=r, column=1).number_format = "YYYY-MM-DD"
            ws_cardio.cell(row=r, column=2, value=we.session.name or "").border = thin_border
            ws_cardio.cell(row=r, column=3, value=we.exercise.name).border = thin_border
            ws_cardio.cell(row=r, column=4, value=we.exercise.category).border = thin_border

            try:
                cd = we.cardio_details
                ws_cardio.cell(row=r, column=5, value=cd.duration_minutes or "").border = thin_border
                c = ws_cardio.cell(row=r, column=6, value=float(cd.distance) if cd.distance else "")
                c.border = thin_border
                ws_cardio.cell(row=r, column=7, value=cd.intensity or "").border = thin_border
                ws_cardio.cell(row=r, column=8, value=cd.calories_burned or "").border = thin_border
                ws_cardio.cell(row=r, column=9, value=cd.avg_heart_rate or "").border = thin_border
            except CardioDetails.DoesNotExist:
                try:
                    cld = we.class_details
                    ws_cardio.cell(row=r, column=5, value=cld.duration_minutes or "").border = thin_border
                    ws_cardio.cell(row=r, column=7, value=cld.intensity or "").border = thin_border
                except ClassDetails.DoesNotExist:
                    pass
                for c in range(5, 10):
                    if not ws_cardio.cell(row=r, column=c).value:
                        ws_cardio.cell(row=r, column=c).border = thin_border

            r += 1

        for col in range(1, len(headers) + 1):
            ws_cardio.column_dimensions[get_column_letter(col)].width = max(14, len(headers[col - 1]) + 4)

        # ── Sheet 5: Personal Records ──
        ws_pr = wb.create_sheet("Personal Records")
        ws_pr.sheet_properties.tabColor = "FFD700"

        headers = ["Exercise", "Weight (lbs)", "Reps", "Est. 1RM", "Date Achieved"]
        for col_idx, h in enumerate(headers, start=1):
            ws_pr.cell(row=1, column=col_idx, value=h)
        style_header_row(ws_pr, 1, len(headers))

        r = 2
        for pr in prs:
            ws_pr.cell(row=r, column=1, value=pr.exercise.name).border = thin_border
            ws_pr.cell(row=r, column=2, value=float(pr.weight)).border = thin_border
            ws_pr.cell(row=r, column=3, value=pr.reps).border = thin_border
            c = ws_pr.cell(row=r, column=4, value=round(float(pr.estimated_1rm), 1) if pr.estimated_1rm else "")
            c.border = thin_border
            ws_pr.cell(row=r, column=5, value=pr.achieved_date).border = thin_border
            ws_pr.cell(row=r, column=5).number_format = "YYYY-MM-DD"
            r += 1

        for col in range(1, len(headers) + 1):
            ws_pr.column_dimensions[get_column_letter(col)].width = max(14, len(headers[col - 1]) + 4)

        # ── Sheet 6: Steps & Activity ──
        ws_steps = wb.create_sheet("Steps & Activity")
        ws_steps.sheet_properties.tabColor = "00BCD4"

        headers = ["Date", "Steps", "Goal", "Distance (mi)", "Calories", "Flights Climbed", "Exercise Min", "Stand Hours"]
        for col_idx, h in enumerate(headers, start=1):
            ws_steps.cell(row=1, column=col_idx, value=h)
        style_header_row(ws_steps, 1, len(headers))

        r = 2
        for s in steps:
            ws_steps.cell(row=r, column=1, value=s.logged_date).border = thin_border
            ws_steps.cell(row=r, column=1).number_format = "YYYY-MM-DD"
            ws_steps.cell(row=r, column=2, value=s.count).border = thin_border
            ws_steps.cell(row=r, column=2).number_format = "#,##0"
            ws_steps.cell(row=r, column=3, value=s.goal or "").border = thin_border
            c = ws_steps.cell(row=r, column=4, value=float(s.distance_miles) if s.distance_miles else "")
            c.border = thin_border
            ws_steps.cell(row=r, column=5, value=s.calories_burned or "").border = thin_border
            ws_steps.cell(row=r, column=6, value=s.flights_climbed or "").border = thin_border
            ws_steps.cell(row=r, column=7, value=s.exercise_minutes or "").border = thin_border
            ws_steps.cell(row=r, column=8, value=s.stand_hours or "").border = thin_border
            r += 1

        # Steps chart (last 30 days)
        if steps.exists():
            steps_chart_start = 2
            steps_chart_end = min(r - 1, 32)  # up to 30 days
            if steps_chart_end >= steps_chart_start:
                chart3 = BarChart()
                chart3.title = "Daily Steps"
                chart3.y_axis.title = "Steps"
                chart3.style = 10
                data = Reference(ws_steps, min_col=2, min_row=1, max_row=steps_chart_end)
                cats = Reference(ws_steps, min_col=1, min_row=steps_chart_start, max_row=steps_chart_end)
                chart3.add_data(data, titles_from_data=True)
                chart3.set_categories(cats)
                ws_steps.add_chart(chart3, f"J2")

        for col in range(1, len(headers) + 1):
            ws_steps.column_dimensions[get_column_letter(col)].width = max(14, len(headers[col - 1]) + 4)

        # ── Sheet 7: Exercise Progress ──
        ws_prog = wb.create_sheet("Exercise Progress")
        ws_prog.sheet_properties.tabColor = "9C27B0"

        # For each exercise, show best set per workout over time
        exercise_workouts = defaultdict(list)
        for we in workout_exercises:
            best_set = we.sets.order_by("-weight").first()
            if best_set and best_set.weight:
                exercise_workouts[we.exercise.name].append(
                    {
                        "date": we.session.date,
                        "weight": float(best_set.weight),
                        "reps": best_set.reps or 0,
                        "volume": float((best_set.weight or 0) * (best_set.reps or 0)),
                    }
                )

        row = 1
        ws_prog.cell(row=row, column=1, value="Exercise Progress — Best Set Per Workout").font = Font(
            bold=True, size=14, color="2B5797"
        )
        ws_prog.merge_cells("A1:E1")
        row += 2

        for exercise_name, sessions in sorted(exercise_workouts.items()):
            if len(sessions) < 1:
                continue
            ws_prog.cell(row=row, column=1, value=exercise_name).font = subheader_font
            ws_prog.cell(row=row, column=1).fill = subheader_fill
            for c in range(1, 5):
                ws_prog.cell(row=row, column=c).fill = subheader_fill
            row += 1

            headers_p = ["Date", "Weight (lbs)", "Reps", "Volume"]
            for col_idx, h in enumerate(headers_p, start=1):
                ws_prog.cell(row=row, column=col_idx, value=h)
            style_header_row(ws_prog, row, 4)
            row += 1

            for session in sorted(sessions, key=lambda x: x["date"]):
                ws_prog.cell(row=row, column=1, value=session["date"]).border = thin_border
                ws_prog.cell(row=row, column=1).number_format = "YYYY-MM-DD"
                ws_prog.cell(row=row, column=2, value=session["weight"]).border = thin_border
                ws_prog.cell(row=row, column=3, value=session["reps"]).border = thin_border
                c = ws_prog.cell(row=row, column=4, value=session["volume"])
                c.border = thin_border
                c.number_format = "#,##0"
                row += 1

            row += 1  # Blank row between exercises

        for col, width in [(1, 16), (2, 16), (3, 10), (4, 14), (5, 14)]:
            ws_prog.column_dimensions[get_column_letter(col)].width = width

        # Write to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"workout_dashboard_{today.strftime('%Y-%m-%d')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
